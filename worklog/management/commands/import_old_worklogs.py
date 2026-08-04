import openpyxl
from datetime import datetime, time
from decimal import Decimal
from django.core.management.base import BaseCommand
from django.contrib.auth.models import User
from worklog.models import StaffProfile, WorkLog

class Command(BaseCommand):
    help = 'Import old worklogs from Excel file'

    def add_arguments(self, parser):
        parser.add_argument('excel_file', type=str, help='Path to Excel file')
        parser.add_argument('--dry-run', action='store_true', help='Preview without saving')

    def handle(self, *args, **options):
        excel_file = options['excel_file']
        dry_run = options['dry_run']

        staff_mapping = {
            "AKABUEZE C.": self.get_staff_by_id(37),
            "ADJATOR E.": self.get_staff_by_id(44),
            "ABRAHAM K.": self.get_staff_by_id(48),
            "ADENIKE O.": self.get_staff_by_id(32),
            "BALOGUN J.": self.get_staff_by_id(43),
            "FOLARIN": self.get_staff_by_id(41),
            "ISHIE N.": self.get_staff_by_id(46),
            "EWA S.": self.get_staff_by_id(28),
            "FAVOUR A.": self.get_staff_by_id(33),
            "KENNENTH E.": self.get_staff_by_id(47),
            "AIDHA I.": self.get_staff_by_id(36),
            "OGBONNA B.": self.get_staff_by_id(39),
            "CHIDERA I.": self.get_staff_by_id(35),
            "DANIEL O.": self.get_staff_by_id(42),
            "EMMANUELLA I.": self.get_staff_by_id(40),
            "NAOMI C.": self.get_staff_by_id(30),
            "OLAMILEKAN": self.get_staff_by_id(49),
            "OGUNMOROTI O.": self.get_staff_by_id(34),
            "LENYIE P.": self.get_staff_by_id(38),
            "WINFIRED": self.get_staff_by_id(45),
        }

        self.stdout.write(f"Loading Excel file: {excel_file}")
        
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Error loading file: {e}"))
            return

        staff_headers = []
        staff_columns = {}
        
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=2, column=col).value
            if cell_value and isinstance(cell_value, str):
                for name in staff_mapping.keys():
                    if name in cell_value:
                        staff_headers.append(cell_value)
                        staff_columns[col] = cell_value
                        break

        self.stdout.write(f"Found {len(staff_headers)} staff columns")

        if not staff_headers:
            self.stdout.write(self.style.ERROR("No staff headers found"))
            return

        current_date = None
        current_description = None
        entries_to_create = []
        skipped_count = 0

        for row_idx in range(3, ws.max_row + 1):
            row = ws[row_idx]
            first_cell = row[0].value
            
            if first_cell and isinstance(first_cell, datetime):
                current_date = first_cell.date()
                self.stdout.write(f"Found date: {current_date}")
                current_description = None
                continue
            
            if first_cell and isinstance(first_cell, str) and first_cell.strip():
                current_description = first_cell.strip()
                if "DO NOT CUT" in current_description.upper():
                    current_description = None
                    continue
            
            if current_date and current_description:
                for col_idx, staff_name in staff_columns.items():
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell_value = cell.value
                    
                    if cell_value == 1 or (isinstance(cell_value, (int, float)) and cell_value == 1):
                        staff = staff_mapping.get(staff_name)
                        if not staff:
                            skipped_count += 1
                            continue
                        
                        entries_to_create.append({
                            'staff': staff,
                            'date': current_date,
                            'description': current_description,
                            'staff_name': staff_name
                        })
        
        self.stdout.write(f"\nFound {len(entries_to_create)} hour entries to import")
        
        if dry_run:
            self.stdout.write("\n--- DRY RUN ---")
            for entry in entries_to_create[:20]:
                self.stdout.write(f"  {entry['staff_name']} - {entry['date']} - {entry['description'][:50]}")
            self.stdout.write(f"\nTotal: {len(entries_to_create)} entries")
            return
        
        self.stdout.write("\nCreating entries...")
        from collections import defaultdict
        
        grouped_entries = defaultdict(list)
        for entry in entries_to_create:
            key = (entry['staff'].id, entry['date'])
            grouped_entries[key].append(entry)
        
        created_count = 0
        
        for (staff_id, date), entries in grouped_entries.items():
            staff = entries[0]['staff']
            current_hour = 7
            
            for entry in entries:
                if current_hour >= 22:
                    break
                
                start_time = time(hour=current_hour, minute=0)
                end_time = time(hour=current_hour + 1, minute=0)
                
                try:
                    WorkLog.objects.create(
                        staff=staff,
                        date=date,
                        description=entry['description'],
                        hours=Decimal('1.00'),
                        start_time=start_time,
                        end_time=end_time,
                        status='completed',
                        is_locked=True,
                        created_by=staff.user
                    )
                    created_count += 1
                except Exception as e:
                    self.stdout.write(self.style.ERROR(f"Error: {e}"))
                
                current_hour += 1
        
        self.stdout.write(self.style.SUCCESS(f"\nCreated: {created_count} entries"))

    def get_staff_by_id(self, user_id):
        try:
            user = User.objects.get(id=user_id)
            return StaffProfile.objects.get(user=user)
        except (User.DoesNotExist, StaffProfile.DoesNotExist):
            return None
